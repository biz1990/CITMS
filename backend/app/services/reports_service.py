"""
app/services/reports_service.py — Module 8: Reports & Export Service.

Implements:
    • get_ram_inventory()       — RAM per slot/device/dept, JSONB extraction
    • get_software_usage()      — Software install counting, violation detection
    • get_component_summary()   — CPU+RAM+MAINBOARD pivot per device
    • export_excel()            — openpyxl workbook bytes
    • export_pdf()              — reportlab PDF bytes

All three aggregate reports are backed by a 15-minute Redis cache.
"""
from __future__ import annotations

import io
import logging
from datetime import datetime
from typing import Any

from sqlalchemy import Float, Integer, and_, case, cast, exists, func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.cache import cache
from app.models.device import Device, DeviceComponent
from app.models.location import Location
from app.models.software import SoftwareBlacklist, SoftwareInstallation, SoftwareLicense
from app.models.user import Department, User
from app.models.ticket import Ticket
from app.models.procurement import SparePartsInventory
from app.schemas.reports import (
    ComponentFilter,
    ComponentSummaryReport,
    ComponentSummaryRow,
    RAMComponentRow,
    RAMInventoryFilter,
    RAMInventoryReport,
    SoftwareInstallationRow,
    SoftwareSummaryEntry,
    SoftwareUsageFilter,
    SoftwareUsageReport,
    AssetOverviewFilter,
    AssetOverviewReport,
    AssetOverviewEntry,
    InventoryStatusReport,
    InventoryStatusRow,
    TicketSLAReport,
    TicketSLARow,
    RepairCostFilter,
    AssetDepreciationReport,
    AssetDepreciationRow,
    ProcurementFilter
)
from datetime import timedelta

logger = logging.getLogger(__name__)

# Cache TTL for each report type (seconds)
_TTL = {
    "ram_inventory":    900,   # 15 min
    "software_usage":   900,   # 15 min
    "component_summary": 900,  # 15 min
}


class ReportsService:
    """Business logic for all Module 8 reports."""

    def __init__(self, db: AsyncSession) -> None:
        self.db = db

    # ====================================================================== #
    # PUBLIC: get_ram_inventory
    # ====================================================================== #
    async def get_ram_inventory(
        self, filters: RAMInventoryFilter
    ) -> RAMInventoryReport:
        cache_key = cache.make_key("ram_inventory", filters.model_dump())
        cached = await cache.get(cache_key)
        if cached:
            logger.debug("Cache HIT ram_inventory key=%s", cache_key)
            return RAMInventoryReport(**cached)

        rows = await self._query_ram_inventory(filters)

        # ── Python-side aggregation ────────────────────────────────────────
        total_gb        : float            = 0.0
        by_manufacturer : dict[str, float] = {}
        by_department   : dict[str, float] = {}
        by_type         : dict[str, int]   = {}
        by_location     : dict[str, float] = {}

        items: list[RAMComponentRow] = []
        for r in rows:
            cap   = float(r["capacity_gb"] or 0)
            mfr   = r["manufacturer"] or "Unknown"
            dept  = r["department"]   or "Unassigned"
            mtype = r["memory_type"]  or "Unknown"
            loc   = r["location_name"] or "Unassigned"

            total_gb                      += cap
            by_manufacturer[mfr]           = by_manufacturer.get(mfr, 0.0) + cap
            by_department[dept]            = by_department.get(dept, 0.0)   + cap
            by_type[mtype]                 = by_type.get(mtype, 0) + 1
            by_location[loc]               = by_location.get(loc, 0.0)      + cap

            items.append(RAMComponentRow(
                device_id=r["device_id"],
                hostname=r["hostname"],
                asset_tag=r["asset_tag"],
                device_name=r["device_name"],
                location_name=r["location_name"],
                department=r["department"],
                component_id=r["component_id"],
                slot_name=r["slot_name"],
                manufacturer=r["manufacturer"],
                ram_model=r["ram_model"],
                component_serial=r["component_serial"],
                component_status=r["component_status"],
                installation_date=r["installation_date"],
                capacity_gb=r["capacity_gb"],
                memory_type=r["memory_type"],
                speed_mhz=r["speed_mhz"],
                form_factor=r["form_factor"],
            ))

        report = RAMInventoryReport(
            generated_at=datetime.utcnow(),
            total_records=len(items),
            total_gb=round(total_gb, 2),
            by_manufacturer={k: round(v, 2) for k, v in by_manufacturer.items()},
            by_department={k: round(v, 2) for k, v in by_department.items()},
            by_type=by_type,
            by_location={k: round(v, 2) for k, v in by_location.items()},
            items=items,
        )

        await cache.set(cache_key, report.model_dump(), ttl=_TTL["ram_inventory"])
        return report

    # ====================================================================== #
    # PUBLIC: get_software_usage
    # ====================================================================== #
    async def get_software_usage(
        self, filters: SoftwareUsageFilter
    ) -> SoftwareUsageReport:
        cache_key = cache.make_key("software_usage", filters.model_dump())
        cached = await cache.get(cache_key)
        if cached:
            logger.debug("Cache HIT software_usage key=%s", cache_key)
            return SoftwareUsageReport(**cached)

        rows = await self._query_software_usage(filters)

        # ── Python-side aggregation ────────────────────────────────────────
        by_software_dict: dict[str, dict[str, Any]] = {}
        by_department:    dict[str, int]             = {}
        violations:       list[SoftwareInstallationRow] = []
        items:            list[SoftwareInstallationRow] = []

        for r in rows:
            sw   = r["software_name"] or "Unknown"
            ver  = r["version"]       or "N/A"
            dept = r["department"]    or "Unassigned"

            if sw not in by_software_dict:
                by_software_dict[sw] = {"total": 0, "by_version": {}, "by_dept": {}}
            by_software_dict[sw]["total"]            += 1
            by_software_dict[sw]["by_version"][ver]   = by_software_dict[sw]["by_version"].get(ver, 0) + 1
            by_software_dict[sw]["by_dept"][dept]     = by_software_dict[sw]["by_dept"].get(dept, 0) + 1
            by_department[dept]                       = by_department.get(dept, 0) + 1

            row_obj = SoftwareInstallationRow(
                installation_id=r["installation_id"],
                software_name=r["software_name"],
                version=r["version"],
                publisher=r["publisher"],
                install_date=r["install_date"],
                hostname=r["hostname"],
                asset_tag=r["asset_tag"],
                device_name=r["device_name"],
                department=r["department"],
                department_id=r["department_id"],
                license_id=r["license_id"],
                license_type=r["license_type"],
                total_seats=r["total_seats"],
                used_seats=r["used_seats"],
                license_expire_date=r["license_expire_date"],
                over_license=bool(r["over_license"]),
                is_blacklisted=bool(r["is_blacklisted"]),
                is_blocked=bool(r["is_blocked"] or False),
            )
            items.append(row_obj)
            if row_obj.over_license or row_obj.is_blacklisted or row_obj.is_blocked:
                violations.append(row_obj)

        # Build sorted summary list
        by_software = [
            SoftwareSummaryEntry(
                software_name=sw,
                total_installs=data["total"],
                by_version=data["by_version"],
                by_dept=data["by_dept"],
            )
            for sw, data in sorted(
                by_software_dict.items(), key=lambda x: x[1]["total"], reverse=True
            )
        ]

        report = SoftwareUsageReport(
            generated_at=datetime.utcnow(),
            total_records=len(items),
            violation_count=len(violations),
            by_software=by_software,
            by_department=by_department,
            violations=violations,
            items=items,
        )

        await cache.set(cache_key, report.model_dump(), ttl=_TTL["software_usage"])
        return report

    # ====================================================================== #
    # PUBLIC: get_component_summary
    # ====================================================================== #
    async def get_component_summary(
        self, filters: ComponentFilter
    ) -> ComponentSummaryReport:
        cache_key = cache.make_key("component_summary", {
            "dept": str(filters.department_id),
            "loc": str(filters.location_id),
            "types": ",".join(sorted(filters.component_types)),
        })
        cached = await cache.get(cache_key)
        if cached:
            logger.debug("Cache HIT component_summary key=%s", cache_key)
            return ComponentSummaryReport(**cached)

        rows = await self._query_component_summary(filters)

        by_cpu_mfr  : dict[str, int]   = {}
        by_ram_size : dict[str, int]   = {}
        by_mb_mfr   : dict[str, int]   = {}
        items       : list[ComponentSummaryRow] = []

        for r in rows:
            cpu_mfr = r["cpu_manufacturer"] or "Unknown"
            mb_mfr  = r["mainboard_manufacturer"] or "Unknown"
            ram_gb  = r["total_ram_gb"]

            by_cpu_mfr[cpu_mfr] = by_cpu_mfr.get(cpu_mfr, 0) + 1
            by_mb_mfr[mb_mfr]   = by_mb_mfr.get(mb_mfr, 0) + 1
            if ram_gb is not None:
                label = f"{int(ram_gb)}GB"
                by_ram_size[label] = by_ram_size.get(label, 0) + 1

            items.append(ComponentSummaryRow(
                device_id=r["device_id"],
                hostname=r["hostname"],
                asset_tag=r["asset_tag"],
                device_name=r["device_name"],
                department=r["department"],
                os_name=r["os_name"],
                cpu_manufacturer=r["cpu_manufacturer"],
                cpu_model=r["cpu_model"],
                cpu_cores=r["cpu_cores"],
                cpu_threads=r["cpu_threads"],
                cpu_speed_ghz=r["cpu_speed_ghz"],
                total_ram_gb=r["total_ram_gb"],
                ram_slots_used=r["ram_slots_used"] or 0,
                mainboard_manufacturer=r["mainboard_manufacturer"],
                mainboard_model=r["mainboard_model"],
                bios_version=r["bios_version"],
                bios_release_date=r["bios_release_date"],
            ))

        report = ComponentSummaryReport(
            generated_at=datetime.utcnow(),
            total_devices=len(items),
            by_cpu_manufacturer=by_cpu_mfr,
            by_ram_size=by_ram_size,
            by_mainboard_manufacturer=by_mb_mfr,
            items=items,
        )

        await cache.set(cache_key, report.model_dump(), ttl=_TTL["component_summary"])
        return report

    # ====================================================================== #
    # PUBLIC: get_asset_overview (Module 8.1.1)
    # ====================================================================== #
    async def get_asset_overview(self, filters: AssetOverviewFilter) -> AssetOverviewReport:
        # Group by device_type and status
        stmt = select(
            Device.device_type,
            Device.status,
            func.count(Device.id).label("count")
        ).where(Device.deleted_at.is_(None)).group_by(Device.device_type, Device.status)
        
        if filters.location_id:
            stmt = stmt.where(Device.location_id == filters.location_id)
            
        results = (await self.db.execute(stmt)).all()
        
        overview_dict = {}
        total_assets = 0
        for dtype, status, count in results:
            if dtype not in overview_dict:
                overview_dict[dtype] = {"total": 0, "status": {}}
            overview_dict[dtype]["total"] += count
            overview_dict[dtype]["status"][status] = count
            total_assets += count
            
        # Group by location
        loc_stmt = select(Location.name, func.count(Device.id)).join(Device).group_by(Location.name)
        loc_results = (await self.db.execute(loc_stmt)).all()
        
        return AssetOverviewReport(
            generated_at=datetime.utcnow(),
            total_assets=total_assets,
            overview=[
                AssetOverviewEntry(device_type=k, count=v["total"], by_status=v["status"])
                for k, v in overview_dict.items()
            ],
            by_location={name: count for name, count in loc_results}
        )

    # ====================================================================== #
    # PUBLIC: get_inventory_status (Module 8.1.3)
    # ====================================================================== #
    async def get_inventory_status(self) -> InventoryStatusReport:
        stmt = select(SparePartsInventory).where(SparePartsInventory.deleted_at.is_(None))
        items = (await self.db.execute(stmt)).scalars().all()
        
        total_value = 0
        low_stock_count = 0
        rows = []
        for item in items:
            value = float(item.quantity * item.unit_cost)
            total_value += value
            is_low = item.quantity <= item.min_quantity
            if is_low:
                low_stock_count += 1
                
            rows.append(InventoryStatusRow(
                name=item.name,
                category=item.category or "Unknown",
                quantity=item.quantity,
                min_quantity=item.min_quantity,
                unit_cost=float(item.unit_cost),
                total_value=value,
                is_low_stock=is_low
            ))
            
        return InventoryStatusReport(
            generated_at=datetime.utcnow(),
            total_value=total_value,
            low_stock_count=low_stock_count,
            items=rows
        )

    # ====================================================================== #
    # PUBLIC: get_offline_devices (Module 8.1.4)
    # ====================================================================== #
    async def get_offline_devices(self, days: int = 7) -> list[dict]:
        threshold = datetime.utcnow() - timedelta(days=days)
        stmt = select(Device).where(
            Device.last_seen < threshold,
            Device.status == "IN_USE",
            Device.deleted_at.is_(None)
        ).order_by(Device.last_seen.asc())
        
        results = (await self.db.execute(stmt)).scalars().all()
        return [
            {
                "hostname": d.hostname,
                "asset_tag": d.asset_tag,
                "last_seen": d.last_seen,
                "days_offline": (datetime.utcnow() - d.last_seen).days if d.last_seen else 999
            }
            for d in results
        ]

    # ====================================================================== #
    # PUBLIC: get_license_expiring (Module 8.1.7)
    # ====================================================================== #
    async def get_license_expiring(self, days: int = 30) -> list[dict]:
        threshold = datetime.utcnow().date() + timedelta(days=days)
        stmt = select(SoftwareLicense).where(
            SoftwareLicense.expire_date <= threshold,
            SoftwareLicense.expire_date >= datetime.utcnow().date(),
            SoftwareLicense.deleted_at.is_(None)
        ).order_by(SoftwareLicense.expire_date.asc())
        
        results = (await self.db.execute(stmt)).scalars().all()
        return [
            {
                "software_name": l.software_name,
                "expire_date": l.expire_date,
                "days_remaining": (l.expire_date - datetime.utcnow().date()).days if l.expire_date else 0,
                "total_seats": l.total_seats,
                "used_seats": l.used_seats
            }
            for l in results
        ]

    # ====================================================================== #
    # PUBLIC: get_ticket_sla_stats (Module 8.1.8) - REFAC: Materialized View
    # ====================================================================== #
    async def get_ticket_sla_stats(self) -> TicketSLAReport:
        """
        v3.6 §3.11: Querying mv_ticket_sla_stats for Enterprise-grade performance.
        """
        from sqlalchemy import text
        stmt = text("SELECT * FROM mv_ticket_sla_stats")
        results = (await self.db.execute(stmt)).mappings().all()

        total = 0
        breached = 0
        total_hours = 0.0
        resolved_count = 0
        by_priority = {}
        rows = []

        for r in results:
            total += 1
            if r["is_breached"]:
                breached += 1
            
            if r["resolution_time_hours"] is not None:
                total_hours += float(r["resolution_time_hours"])
                resolved_count += 1
            
            prio = r["priority"] or "MEDIUM"
            if prio not in by_priority:
                by_priority[prio] = {"count": 0, "breaches": 0}
            by_priority[prio]["count"] += 1
            if r["is_breached"]:
                by_priority[prio]["breaches"] += 1
            
            rows.append(TicketSLARow(
                ticket_code=r["ticket_code"],
                title=r["title"],
                priority=prio,
                created_at=r["created_at"],
                resolved_at=r["resolved_at"],
                sla_resolution_due=r["sla_resolution_due"],
                is_breached=r["is_breached"],
                resolution_time_hours=r["resolution_time_hours"]
            ))

        return TicketSLAReport(
            generated_at=datetime.utcnow(),
            total_tickets=total,
            breach_count=breached,
            avg_resolution_time_hours=total_hours / max(resolved_count, 1),
            by_priority=by_priority,
            items=rows
        )

    # ====================================================================== #
    # PUBLIC: get_repair_costs (Module 8.1.5)
    # ====================================================================== #
    async def get_repair_costs(self, filters: RepairCostFilter) -> dict:
        from app.models.ticket import MaintenanceLog
        stmt = select(
            func.date_trunc('month', MaintenanceLog.created_at).label("month"),
            func.sum(MaintenanceLog.cost).label("total_cost"),
            func.count(MaintenanceLog.id).label("count")
        ).where(MaintenanceLog.deleted_at.is_(None)).group_by("month").order_by("month")
        
        results = (await self.db.execute(stmt)).all()
        return {
            "generated_at": datetime.utcnow(),
            "monthly_costs": [
                {"month": str(r.month), "total_cost": float(r.total_cost), "count": r.count}
                for r in results
            ]
        }

    # ====================================================================== #
    # PUBLIC: get_asset_depreciation (Module 8.1.2) - REFAC: Materialized View
    # ====================================================================== #
    async def get_asset_depreciation(self) -> AssetDepreciationReport:
        """
        v3.6 §3.11: Querying mv_asset_depreciation for pre-calculated TCO.
        """
        from sqlalchemy import text
        stmt = text("SELECT * FROM mv_asset_depreciation")
        results = (await self.db.execute(stmt)).mappings().all()
        
        total_purchase = 0.0
        total_current = 0.0
        rows = []
        
        for r in results:
            cost = float(r["purchase_cost"] or 0)
            curr = float(r["current_value"] or 0)
            total_purchase += cost
            total_current += curr
            
            rows.append(AssetDepreciationRow(
                asset_tag=r["asset_tag"],
                name=r["name"],
                purchase_date=r["purchase_date"],
                purchase_cost=cost,
                current_value=round(curr, 2),
                depreciation_method="STRAIGHT_LINE",
                age_months=r["age_months"]
            ))
            
        return AssetDepreciationReport(
            generated_at=datetime.utcnow(),
            total_purchase_cost=round(total_purchase, 2),
            total_current_value=round(total_current, 2),
            items=rows
        )

    async def refresh_materialized_views(self):
        """
        v3.6 §10: Manual trigger to refresh all MVs. 
        In production, this is handled by pg_cron.
        """
        from sqlalchemy import text
        views = [
            "mv_asset_depreciation",
            "mv_software_usage_top10",
            "mv_ticket_sla_stats",
            "mv_offline_missing_devices"
        ]
        for v in views:
            await self.db.execute(text(f"REFRESH MATERIALIZED VIEW CONCURRENTLY {v}"))
        await self.db.commit()
        logger.info("All Materialized Views refreshed successfully.")

    # ====================================================================== #
    # PUBLIC: get_procurement_costs (Module 8.1.6)
    # ====================================================================== #
    async def get_procurement_costs(self, filters: ProcurementFilter) -> dict:
        from app.models.procurement import PurchaseOrder
        stmt = select(
            func.date_trunc('month', PurchaseOrder.created_at).label("month"),
            func.sum(PurchaseOrder.total_estimated_cost).label("total_spent")
        ).where(
            PurchaseOrder.status == "COMPLETED",
            PurchaseOrder.deleted_at.is_(None)
        ).group_by("month").order_by("month")
        
        results = (await self.db.execute(stmt)).all()
        return {
            "generated_at": datetime.utcnow(),
            "monthly_spending": [
                {"month": str(r.month), "total_spent": float(r.total_spent or 0)}
                for r in results
            ]
        }

    # ====================================================================== #
    # PUBLIC: export_excel
    # ====================================================================== #
    async def export_excel(self, report_type: str, filters: Any) -> bytes:
        """Return an openpyxl workbook as raw bytes."""
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError as exc:
            raise RuntimeError("openpyxl is required for Excel export") from exc

        wb = openpyxl.Workbook()

        if report_type == "ram_inventory":
            report = await self.get_ram_inventory(filters)
            ws_data = wb.active
            ws_data.title = "RAM Inventory"
            ws_data.append([
                "Hostname", "Asset Tag", "Department", "Location",
                "Slot", "Manufacturer", "Model", "Capacity (GB)",
                "Type", "Speed (MHz)", "Form Factor", "Status", "Install Date",
            ])
            for i, item in enumerate(report.items, 2):
                ws_data.append([
                    item.hostname, item.asset_tag, item.department, item.location_name,
                    item.slot_name, item.manufacturer, item.ram_model, item.capacity_gb,
                    item.memory_type, item.speed_mhz, item.form_factor,
                    item.component_status, str(item.installation_date) if item.installation_date else None,
                ])

        elif report_type == "software_usage":
            report = await self.get_software_usage(filters)
            ws_data = wb.active
            ws_data.title = "Software Usage"
            ws_data.append([
                "Software Name", "Version", "Publisher", "Hostname", "Asset Tag",
                "Department", "License Type", "Total Seats", "Used Seats", "Over License"
            ])
            for item in report.items:
                ws_data.append([
                    item.software_name, item.version, item.publisher,
                    item.hostname, item.asset_tag, item.department,
                    item.license_type, item.total_seats, item.used_seats, "YES" if item.over_license else "no"
                ])

        elif report_type == "component_summary":
            report = await self.get_component_summary(filters)
            ws_data = wb.active
            ws_data.title = "Component Summary"
            ws_data.append([
                "Hostname", "Asset Tag", "Department", "OS",
                "CPU Manufacturer", "CPU Model", "Cores", "Total RAM (GB)", "Mainboard"
            ])
            for item in report.items:
                ws_data.append([
                    item.hostname, item.asset_tag, item.department, item.os_name,
                    item.cpu_manufacturer, item.cpu_model, item.cpu_cores, item.total_ram_gb, item.mainboard_model
                ])

        elif report_type == "asset_overview":
            report = await self.get_asset_overview(filters)
            ws = wb.active
            ws.title = "Overview"
            ws.append(["Asset Type", "Total", "Status Breakdown"])
            for e in report.overview:
                ws.append([e.device_type, e.count, str(e.by_status)])

        elif report_type == "depreciation":
            report = await self.get_asset_depreciation()
            ws = wb.active
            ws.title = "TCO & Depreciation"
            headers = ["Asset Tag", "Name", "Purchase Date", "Cost", "Current Value", "Age (Months)"]
            ws.append(headers)
            for i, item in enumerate(report.items, 2):
                ws.append([item.asset_tag, item.name, str(item.purchase_date), item.purchase_cost, item.current_value, item.age_months])

        elif report_type == "inventory_status":
            report = await self.get_inventory_status()
            ws = wb.active
            ws.title = "Inventory Status"
            headers = ["Name", "Category", "Quantity", "Min Qty", "Unit Cost", "Total Value", "Low Stock"]
            ws.append(headers)
            for i, item in enumerate(report.items, 2):
                ws.append([item.name, item.category, item.quantity, item.min_quantity, item.unit_cost, item.total_value, "YES" if item.is_low_stock else "no"])

        elif report_type == "offline_devices":
            items = await self.get_offline_devices()
            ws = wb.active
            ws.title = "Offline Devices"
            ws.append(["Hostname", "Asset Tag", "Last Seen", "Days Offline"])
            for r in items:
                ws.append([r["hostname"], r["asset_tag"], str(r["last_seen"]), r["days_offline"]])

        elif report_type == "repair_costs":
            report = await self.get_repair_costs(filters)
            ws = wb.active
            ws.title = "Monthly Repair Costs"
            ws.append(["Month", "Total Cost", "Ticket Count"])
            for r in report["monthly_costs"]:
                ws.append([r["month"], r["total_cost"], r["count"]])

        elif report_type == "procurement_costs":
            report = await self.get_procurement_costs(filters)
            ws = wb.active
            ws.title = "Monthly Procurement"
            ws.append(["Month", "Total Spent"])
            for r in report["monthly_spending"]:
                ws.append([r["month"], r["total_spent"]])

        elif report_type == "license_expiring":
            items = await self.get_license_expiring()
            ws = wb.active
            ws.title = "License Expiry"
            ws.append(["Software", "Expire Date", "Days Remaining", "Seats Used", "Total Seats"])
            for r in items:
                ws.append([r["software_name"], str(r["expire_date"]), r["days_remaining"], r["used_seats"], r["total_seats"]])

        elif report_type == "ticket_sla":
            report = await self.get_ticket_sla_stats()
            ws = wb.active
            ws.title = "Ticket SLA"
            ws.append(["Ticket Code", "Title", "Priority", "Created At", "Resolved At", "Breached", "Resolution Hours"])
            for r in report.items:
                ws.append([r.ticket_code, r.title, r.priority, str(r.created_at), str(r.resolved_at), "YES" if r.is_breached else "no", r.resolution_time_hours])

        # ── Auto-width and Stylizing (Shared) ──────────────────────────────
        for ws in wb.worksheets:
            for col in ws.columns:
                max_len = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_len:
                            max_len = len(str(cell.value))
                    except: pass
                adjusted_width = (max_len + 2)
                ws.column_dimensions[column].width = min(adjusted_width, 50)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read()

    # ====================================================================== #
    # PUBLIC: export_pdf
    # ====================================================================== #
    async def export_pdf(self, report_type: str, filters: Any) -> bytes:
        """Return a reportlab PDF as raw bytes."""
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib.units import cm
            from reportlab.platypus import (
                Paragraph,
                SimpleDocTemplate,
                Spacer,
                Table,
                TableStyle,
            )
        except ImportError as exc:
            raise RuntimeError("reportlab is required for PDF export") from exc

        buf = io.BytesIO()
        styles  = getSampleStyleSheet()
        title_s = styles["Title"]
        h2_s    = styles["Heading2"]
        normal_s = styles["Normal"]

        # Shared table style
        def _base_table_style(header_color: colors.Color) -> TableStyle:
            return TableStyle([
                ("BACKGROUND",   (0, 0), (-1, 0),  header_color),
                ("TEXTCOLOR",    (0, 0), (-1, 0),  colors.white),
                ("FONTNAME",     (0, 0), (-1, 0),  "Helvetica-Bold"),
                ("FONTSIZE",     (0, 0), (-1, 0),  9),
                ("ALIGN",        (0, 0), (-1, 0),  "CENTER"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
                ("FONTSIZE",     (0, 1), (-1, -1), 8),
                ("GRID",         (0, 0), (-1, -1), 0.25, colors.grey),
                ("VALIGN",       (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING",  (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING",   (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING",(0, 0), (-1, -1), 3),
            ])

        elements: list = []

        if report_type == "ram_inventory":
            report = await self.get_ram_inventory(filters)
            doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=1*cm, rightMargin=1*cm)
            elements += [
                Paragraph("RAM Inventory Report", title_s),
                Paragraph(f"Generated: {report.generated_at.strftime('%Y-%m-%d %H:%M UTC')}  |  "
                          f"Total: {report.total_records} modules  |  "
                          f"Total RAM: {report.total_gb} GB", normal_s),
                Spacer(1, 0.4*cm),
                Paragraph("Summary by Manufacturer", h2_s),
            ]
            mfr_data = [["Manufacturer", "Total GB"]] + [
                [k, f"{v:.1f}"] for k, v in sorted(report.by_manufacturer.items(), key=lambda x: x[1], reverse=True)
            ]
            elements += [
                Table(mfr_data, colWidths=[8*cm, 4*cm],
                      style=_base_table_style(colors.HexColor("#1E40AF"))),
                Spacer(1, 0.5*cm),
                Paragraph("Detail — Per RAM Module", h2_s),
            ]
            detail_data = [["Hostname", "Asset Tag", "Dept", "Slot", "Mfr", "Capacity", "Type", "MHz"]] + [
                [
                    item.hostname or "", item.asset_tag or "", item.department or "",
                    item.slot_name or "", item.manufacturer or "",
                    f"{item.capacity_gb or 0:.0f} GB", item.memory_type or "", str(item.speed_mhz or ""),
                ]
                for item in report.items
            ]
            col_w = [4*cm, 3*cm, 3.5*cm, 2.5*cm, 3*cm, 2.5*cm, 2*cm, 2*cm]
            elements.append(Table(detail_data, colWidths=col_w,
                                   style=_base_table_style(colors.HexColor("#1E40AF"))))

        elif report_type == "software_usage":
            report = await self.get_software_usage(filters)
            doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=1*cm, rightMargin=1*cm)

            vio_style = _base_table_style(colors.HexColor("#065F46"))
            vio_style.add("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#065F46"))

            elements += [
                Paragraph("Software Usage Report", title_s),
                Paragraph(f"Generated: {report.generated_at.strftime('%Y-%m-%d %H:%M UTC')}  |  "
                          f"Total: {report.total_records} installations  |  "
                          f"Violations: {report.violation_count}", normal_s),
                Spacer(1, 0.4*cm),
                Paragraph("Top Software by Machine Count", h2_s),
            ]
            top10 = report.by_software[:10]
            sw_data = [["Software Name", "# Machines"]] + [[e.software_name, e.total_installs] for e in top10]
            elements += [
                Table(sw_data, colWidths=[12*cm, 4*cm],
                      style=_base_table_style(colors.HexColor("#065F46"))),
                Spacer(1, 0.5*cm),
            ]
            if report.violations:
                elements.append(Paragraph("⚠ License Violations & Blacklisted Software", h2_s))
                vio_data = [["Software", "Version", "Hostname", "Dept", "Over License", "Blacklisted"]] + [
                    [
                        v.software_name, v.version or "", v.hostname or "", v.department or "",
                        "YES" if v.over_license else "no",
                        "YES" if v.is_blacklisted else "no",
                    ]
                    for v in report.violations
                ]
                vio_ts = _base_table_style(colors.HexColor("#991B1B"))
                # Highlight over_license rows red
                for i, v in enumerate(report.violations, 1):
                    if v.is_blacklisted:
                        vio_ts.add("BACKGROUND", (0, i), (-1, i), colors.HexColor("#FEF3C7"))
                    elif v.over_license:
                        vio_ts.add("BACKGROUND", (0, i), (-1, i), colors.HexColor("#FEE2E2"))
                elements.append(Table(vio_data, colWidths=[5*cm, 3*cm, 4*cm, 3*cm, 3*cm, 3*cm], style=vio_ts))

        elif report_type == "component_summary":
            report = await self.get_component_summary(filters)
            doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=1*cm, rightMargin=1*cm)
            elements += [
                Paragraph("Device Component Summary", title_s),
                Paragraph(f"Generated: {report.generated_at.strftime('%Y-%m-%d %H:%M UTC')}  |  "
                          f"Total Devices: {report.total_devices}", normal_s),
                Spacer(1, 0.4*cm),
                Paragraph("Device Hardware Details", h2_s),
            ]
            detail_data = [["Hostname", "Asset Tag", "Dept", "CPU", "Cores", "RAM (GB)", "Mainboard", "BIOS"]] + [
                [
                    item.hostname or "", item.asset_tag or "", item.department or "",
                    f"{item.cpu_manufacturer or ''} {item.cpu_model or ''}".strip(),
                    str(item.cpu_cores or ""), f"{item.total_ram_gb or 0:.0f}",
                    f"{item.mainboard_manufacturer or ''} {item.mainboard_model or ''}".strip(),
                    item.bios_version or "",
                ]
                for item in report.items
            ]
            col_w = [4*cm, 3*cm, 3.5*cm, 5*cm, 2*cm, 2.5*cm, 5*cm, 3*cm]
            elements.append(Table(detail_data, colWidths=col_w,
                                   style=_base_table_style(colors.HexColor("#4C0519"))))
        elif report_type == "asset_overview":
            report = await self.get_asset_overview(filters)
            doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=1*cm, rightMargin=1*cm)
            elements += [
                Paragraph("Asset Overview Report", title_s),
                Paragraph(f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}", normal_s),
                Paragraph(f"Total Assets: {report.total_assets}", h2_s),
                Spacer(1, 0.4*cm),
            ]
            ov_data = [["Asset Type", "Count"]] + [[e.device_type, e.count] for e in report.overview]
            elements.append(Table(ov_data, colWidths=[10*cm, 4*cm], style=_base_table_style(colors.grey)))

        elif report_type == "depreciation":
            report = await self.get_asset_depreciation()
            doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=1*cm, rightMargin=1*cm)
            elements += [
                Paragraph("Asset Depreciation (TCO) Report", title_s),
                Paragraph(f"Total Current Value: ${report.total_current_value:,.2f}", normal_s),
                Spacer(1, 0.5*cm),
            ]
            dep_data = [["Tag", "Name", "Cost", "Current Value", "Age (Mo)"]] + [
                [i.asset_tag, i.name, f"${i.purchase_cost:,.2f}", f"${i.current_value:,.2f}", i.age_months]
                for i in report.items
            ]
            elements.append(Table(dep_data, colWidths=[4*cm, 8*cm, 4*cm, 4*cm, 3*cm], style=_base_table_style(colors.darkblue)))

        elif report_type == "inventory_status":
            report = await self.get_inventory_status()
            doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=1*cm, rightMargin=1*cm)
            elements += [
                Paragraph("Inventory Status Report", title_s),
                Paragraph(f"Total Inventory Value: ${report.total_value:,.2f}", normal_s),
                Spacer(1, 0.5*cm),
            ]
            inv_data = [["Item", "Category", "Qty", "Value"]] + [
                [i.name, i.category, i.quantity, f"${i.total_value:,.2f}"]
                for i in report.items
            ]
            elements.append(Table(inv_data, colWidths=[8*cm, 4*cm, 2*cm, 4*cm], style=_base_table_style(colors.darkgreen)))

        else:
            doc = SimpleDocTemplate(buf, pagesize=A4)
            elements.append(Paragraph(f"Report: {report_type}", title_s))
            elements.append(Paragraph("Detailed PDF layout is under construction. Please use Excel for full data.", normal_s))

        doc.build(elements)
        buf.seek(0)
        return buf.read()

    # ====================================================================== #
    # PRIVATE: SQL queries
    # ====================================================================== #
    async def _query_ram_inventory(self, filters: RAMInventoryFilter):
        """Execute the RAM inventory join query and return raw mappings."""
        stmt = (
            select(
                Device.id.label("device_id"),
                Device.hostname,
                Device.asset_tag,
                Device.name.label("device_name"),
                Location.name.label("location_name"),
                Department.name.label("department"),
                DeviceComponent.id.label("component_id"),
                DeviceComponent.slot_name,
                DeviceComponent.manufacturer,
                DeviceComponent.model.label("ram_model"),
                DeviceComponent.serial_number.label("component_serial"),
                DeviceComponent.status.label("component_status"),
                DeviceComponent.installation_date,
                # JSONB extractions (→> returns text; cast to target type)
                cast(DeviceComponent.specifications["capacity_gb"].astext, Float).label("capacity_gb"),
                DeviceComponent.specifications["memory_type"].astext.label("memory_type"),
                cast(DeviceComponent.specifications["speed_mhz"].astext, Integer).label("speed_mhz"),
                DeviceComponent.specifications["form_factor"].astext.label("form_factor"),
            )
            .select_from(DeviceComponent)
            .join(Device, DeviceComponent.device_id == Device.id)
            .outerjoin(Location, Device.location_id == Location.id)
            .outerjoin(User, Device.assigned_to_id == User.id)
            .outerjoin(Department, User.department_id == Department.id)
            .where(
                DeviceComponent.component_type == "RAM",
                DeviceComponent.deleted_at.is_(None),
                Device.deleted_at.is_(None),
            )
        )

        if filters.department_id:
            stmt = stmt.where(Department.id == filters.department_id)
        if filters.location_id:
            stmt = stmt.where(Device.location_id == filters.location_id)
        if filters.manufacturer:
            stmt = stmt.where(DeviceComponent.manufacturer.ilike(f"%{filters.manufacturer}%"))
        if filters.memory_type:
            stmt = stmt.where(
                DeviceComponent.specifications["memory_type"].astext == filters.memory_type
            )
        if filters.date_from:
            stmt = stmt.where(DeviceComponent.installation_date >= filters.date_from)
        if filters.date_to:
            stmt = stmt.where(DeviceComponent.installation_date <= filters.date_to)

        stmt = stmt.order_by(Department.name, Device.hostname, DeviceComponent.slot_name)
        return (await self.db.execute(stmt)).mappings().all()

    async def _query_software_usage(self, filters: SoftwareUsageFilter):
        """Execute the software usage join query and return raw mappings."""
        # Correlated EXISTS subquery for blacklist detection.
        # Checks: does any blacklist entry appear as a substring in the installation's software_name?
        # e.g. blacklist="uTorrent" hits installation="uTorrent 3.5.5"
        blacklist_check = (
            select(SoftwareBlacklist.id)
            .where(
                SoftwareInstallation.software_name.ilike(
                    func.concat("%", SoftwareBlacklist.software_name, "%")
                )
            )
            .correlate(SoftwareInstallation)
            .exists()
            .label("is_blacklisted")
        )

        stmt = (
            select(
                SoftwareInstallation.id.label("installation_id"),
                SoftwareInstallation.software_name,
                SoftwareInstallation.version,
                SoftwareInstallation.publisher,
                SoftwareInstallation.install_date,
                SoftwareInstallation.is_blocked,
                Device.id.label("device_id_col"),
                Device.hostname,
                Device.asset_tag,
                Device.name.label("device_name"),
                Department.id.label("department_id"),
                Department.name.label("department"),
                SoftwareLicense.id.label("license_id"),
                SoftwareLicense.type.label("license_type"),
                SoftwareLicense.total_seats,
                SoftwareLicense.used_seats,
                SoftwareLicense.expire_date.label("license_expire_date"),
                # Inline violation flag (evaluated in SQL, not Python)
                case(
                    (
                        and_(
                            SoftwareLicense.total_seats.is_not(None),
                            SoftwareLicense.used_seats > SoftwareLicense.total_seats,
                        ),
                        True,
                    ),
                    else_=False,
                ).label("over_license"),
                blacklist_check,
            )
            .select_from(SoftwareInstallation)
            .join(Device, SoftwareInstallation.device_id == Device.id)
            .outerjoin(User, Device.assigned_to_id == User.id)
            .outerjoin(Department, User.department_id == Department.id)
            .outerjoin(SoftwareLicense, SoftwareInstallation.license_id == SoftwareLicense.id)
            .where(
                SoftwareInstallation.deleted_at.is_(None),
                Device.deleted_at.is_(None),
            )
        )

        if filters.department_id:
            stmt = stmt.where(Department.id == filters.department_id)
        if filters.software_name:
            stmt = stmt.where(SoftwareInstallation.software_name.ilike(f"%{filters.software_name}%"))
        if filters.license_type:
            stmt = stmt.where(SoftwareLicense.type == filters.license_type)
        if filters.publisher:
            stmt = stmt.where(SoftwareInstallation.publisher.ilike(f"%{filters.publisher}%"))
        if filters.only_violations:
            stmt = stmt.where(
                or_(
                    and_(
                        SoftwareLicense.total_seats.is_not(None),
                        SoftwareLicense.used_seats > SoftwareLicense.total_seats,
                    ),
                    SoftwareInstallation.is_blocked == True,           # noqa: E712
                )
            )

        stmt = stmt.order_by(SoftwareInstallation.software_name, Department.name, Device.hostname)
        return (await self.db.execute(stmt)).mappings().all()

    async def _query_component_summary(self, filters: ComponentFilter):
        """Execute the component pivot query (conditional aggregation)."""
        component_types = filters.component_types or ["CPU", "RAM", "MAINBOARD"]

        stmt = (
            select(
                Device.id.label("device_id"),
                Device.hostname,
                Device.asset_tag,
                Device.name.label("device_name"),
                Device.os_name,
                Department.name.label("department"),
                # ── CPU columns ─────────────────────────────────────────────
                func.max(
                    case((DeviceComponent.component_type == "CPU", DeviceComponent.manufacturer), else_=None)
                ).label("cpu_manufacturer"),
                func.max(
                    case((DeviceComponent.component_type == "CPU", DeviceComponent.model), else_=None)
                ).label("cpu_model"),
                func.max(
                    case(
                        (DeviceComponent.component_type == "CPU",
                         cast(DeviceComponent.specifications["cores"].astext, Integer)),
                        else_=None,
                    )
                ).label("cpu_cores"),
                func.max(
                    case(
                        (DeviceComponent.component_type == "CPU",
                         cast(DeviceComponent.specifications["threads"].astext, Integer)),
                        else_=None,
                    )
                ).label("cpu_threads"),
                func.max(
                    case(
                        (DeviceComponent.component_type == "CPU",
                         cast(DeviceComponent.specifications["speed_ghz"].astext, Float)),
                        else_=None,
                    )
                ).label("cpu_speed_ghz"),
                # ── RAM columns (aggregate across all slots) ─────────────────
                func.sum(
                    case(
                        (DeviceComponent.component_type == "RAM",
                         cast(DeviceComponent.specifications["capacity_gb"].astext, Float)),
                        else_=0.0,
                    )
                ).label("total_ram_gb"),
                func.count(
                    case((DeviceComponent.component_type == "RAM", DeviceComponent.id), else_=None)
                ).label("ram_slots_used"),
                # ── Mainboard columns ────────────────────────────────────────
                func.max(
                    case((DeviceComponent.component_type == "MAINBOARD", DeviceComponent.manufacturer), else_=None)
                ).label("mainboard_manufacturer"),
                func.max(
                    case((DeviceComponent.component_type == "MAINBOARD", DeviceComponent.model), else_=None)
                ).label("mainboard_model"),
                func.max(
                    case(
                        (DeviceComponent.component_type == "MAINBOARD",
                         DeviceComponent.specifications["bios_version"].astext),
                        else_=None,
                    )
                ).label("bios_version"),
                func.max(
                    case(
                        (DeviceComponent.component_type == "MAINBOARD",
                         DeviceComponent.specifications["bios_release_date"].astext),
                        else_=None,
                    )
                ).label("bios_release_date"),
            )
            .select_from(Device)
            .join(DeviceComponent, DeviceComponent.device_id == Device.id)
            .outerjoin(User, Device.assigned_to_id == User.id)
            .outerjoin(Department, User.department_id == Department.id)
            .where(
                Device.deleted_at.is_(None),
                DeviceComponent.deleted_at.is_(None),
                DeviceComponent.component_type.in_(component_types),
            )
            .group_by(
                Device.id,
                Device.hostname,
                Device.asset_tag,
                Device.name,
                Device.os_name,
                Department.name,
            )
            .order_by(Department.name, Device.hostname)
        )

        if filters.department_id:
            stmt = stmt.where(Department.id == filters.department_id)
        if filters.location_id:
            stmt = stmt.where(Device.location_id == filters.location_id)

        return (await self.db.execute(stmt)).mappings().all()
